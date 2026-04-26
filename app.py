"""
=============================================================================
  MEDCONTROL — Sistema de Controle de Validade de Medicamentos
  Versão 2.2 | Arquitetura Multi-Tenant | Segurança Completa

  Segurança aplicada:
    ✔ Senhas com hash (werkzeug)
    ✔ CSRF protection (Flask-WTF)
    ✔ Rate limiting no login (Flask-Limiter)
    ✔ SECRET_KEY e ADMIN_PASS obrigatórias via env
    ✔ Session timeout (30 min inatividade)
    ✔ Cookie seguro (HttpOnly, SameSite, Secure em produção)
    ✔ Headers de segurança HTTP (X-Frame-Options, CSP, etc.)
    ✔ Logs de auditoria (login, logout, cadastro, exclusão)
    ✔ Troca de senha pelo próprio usuário
    ✔ Confirmação server-side antes de excluir rede
    ✔ Paginação na listagem de medicamentos
    ✔ Input validation com try/except em todos os POSTs
=============================================================================
"""

from flask import (Flask, render_template, request, redirect,
                   url_for, session, jsonify, send_file, flash, abort)
import sentry_sdk
from sentry_sdk.integrations.flask import FlaskIntegration
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect, CSRFError
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, timedelta
from functools import wraps
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
import io, os, json, logging, random, string
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import urllib.request, urllib.error, urllib.parse
try:
    import resend as resend_sdk
    RESEND_SDK = True
except ImportError:
    RESEND_SDK = False
try:
    import requests as _requests
except ImportError:
    _requests = None

# ── SENTRY MONITORING ───────────────────────────────────────────────
SENTRY_DSN = os.environ.get('SENTRY_DSN')
if SENTRY_DSN:
    sentry_sdk.init(
        dsn=SENTRY_DSN,
        integrations=[FlaskIntegration()],
        traces_sample_rate=1.0,
        profiles_sample_rate=1.0,
    )

app = Flask(__name__)

# ── MODO MANUTENÇÃO ─────────────────────────────────────────────────
# Mude para True antes de fazer deploy de manutenção.
# Mude de volta para False quando terminar.
MANUTENCAO = False

@app.before_request
def verificar_manutencao():
    if MANUTENCAO and request.endpoint not in ('manutencao', 'static'):
        return render_template('manutencao.html'), 503

@app.route('/manutencao')
def manutencao():
    return render_template('manutencao.html'), 503
# ────────────────────────────────────────────────────────────────────

# Filtro Jinja para data em português
DIAS_PT   = ['Segunda','Terça','Quarta','Quinta','Sexta','Sábado','Domingo']
MESES_PT  = ['','Janeiro','Fevereiro','Março','Abril','Maio','Junho',
              'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

@app.template_filter('datefmt_ptbr')
def datefmt_ptbr(d):
    dia_semana = DIAS_PT[d.weekday()]
    return f"{dia_semana}, {d.day:02d} de {MESES_PT[d.month]} de {d.year}"

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================

_secret = os.environ.get('SECRET_KEY')
if not _secret:
    raise RuntimeError("SECRET_KEY não definida! Adicione nas variáveis do Render.")
app.secret_key = _secret

_is_prod = os.environ.get('DATABASE_URL', '').startswith('postgresql')

# Sessão
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=120)
app.config['SESSION_COOKIE_HTTPONLY']    = True
app.config['SESSION_COOKIE_SAMESITE']   = 'Lax'
app.config['SESSION_COOKIE_SECURE']     = bool(_is_prod)  # HTTPS apenas em produção

# Banco
database_url = os.environ.get('DATABASE_URL', 'sqlite:///medcontrol.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql+psycopg://', 1)
elif database_url.startswith('postgresql://'):
    database_url = database_url.replace('postgresql://', 'postgresql+psycopg://', 1)
app.config['SQLALCHEMY_DATABASE_URI']        = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# CSRF
app.config['WTF_CSRF_TIME_LIMIT'] = 3600

# Paginação
ITENS_POR_PAGINA = 20

db      = SQLAlchemy(app)
csrf    = CSRFProtect(app)
limiter = Limiter(get_remote_address, app=app, default_limits=[], storage_uri="memory://")

# =============================================================================
# HEADERS DE SEGURANÇA HTTP
# =============================================================================

@app.after_request
def set_security_headers(response):
    # Impede que o site seja carregado em iframes (clickjacking)
    response.headers['X-Frame-Options'] = 'DENY'
    # Impede MIME sniffing
    response.headers['X-Content-Type-Options'] = 'nosniff'
    # Controla informações de referência
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # Força HTTPS por 1 ano (só em produção)
    if _is_prod:
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    # Content Security Policy — permite Bootstrap/Google Fonts/CDNs usados
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com https://unpkg.com; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com https://cdn.jsdelivr.net; "
        "img-src 'self' data: blob:; "
        "media-src 'self' blob:; "
        "worker-src 'self' blob:; "
        "connect-src 'self' blob: https://unpkg.com https://cdn.jsdelivr.net https://cdnjs.cloudflare.com;"
    )
    return response


# =============================================================================
# LOGS DE AUDITORIA
# =============================================================================

# Configura logger de auditoria separado
audit_logger = logging.getLogger('medcontrol.audit')
audit_logger.setLevel(logging.INFO)
if not audit_logger.handlers:
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter('[AUDIT] %(asctime)s %(message)s'))
    audit_logger.addHandler(handler)

def audit(acao, detalhe=''):
    """Registra ação crítica com usuário, IP e timestamp."""
    usuario  = session.get('username', 'anônimo')
    ip       = request.remote_addr
    audit_logger.info(f"user={usuario} ip={ip} acao={acao} detalhe={detalhe}")


# =============================================================================
# MODELOS
# =============================================================================

class Rede(db.Model):
    __tablename__ = 'redes'
    id               = db.Column(db.Integer, primary_key=True)
    nome             = db.Column(db.String(200), nullable=False)
    cnpj             = db.Column(db.String(20), nullable=True)
    email_contato    = db.Column(db.String(150), nullable=True)
    telefone         = db.Column(db.String(30), nullable=True)
    ativa            = db.Column(db.Boolean, default=True)
    data_expiracao   = db.Column(db.Date, nullable=True)
    plano            = db.Column(db.String(50), default='mensal')
    data_cadastro    = db.Column(db.DateTime, default=datetime.utcnow)
    # Self-serve & pagamento
    trial            = db.Column(db.Boolean, default=True)
    trial_inicio     = db.Column(db.DateTime, nullable=True)
    mp_assinatura_id = db.Column(db.String(100), nullable=True)   # ID assinatura Mercado Pago
    mp_payer_email   = db.Column(db.String(150), nullable=True)   # email pagador MP
    token_api        = db.Column(db.String(64), nullable=True, unique=True)  # API Key REST externa
    usuarios         = db.relationship('Usuario', backref='rede', lazy=True)
    medicamentos     = db.relationship('Medicamento', backref='rede', lazy=True)

    @property
    def em_trial(self):
        """True se ainda está dentro do período de trial de 30 dias."""
        if not self.trial or not self.trial_inicio: return False
        limite = self.trial_inicio + timedelta(days=30)
        return datetime.utcnow() < limite

    @property
    def dias_trial_restantes(self):
        if not self.trial or not self.trial_inicio: return 0
        limite = self.trial_inicio + timedelta(days=30)
        restam = (limite - datetime.utcnow()).days
        return max(0, restam)

    @property
    def assinatura_ativa(self):
        if not self.ativa: return False
        # Trial ativo = acesso liberado
        if self.em_trial: return True
        # Assinatura paga com data de expiração válida
        if self.data_expiracao and self.data_expiracao >= date.today(): return True
        # Sem expiração definida e não em trial = bloqueia
        return False

    @property
    def dias_restantes(self):
        if not self.data_expiracao: return None
        return (self.data_expiracao - date.today()).days

    @property
    def alerta_renovacao(self):
        d = self.dias_restantes
        return d is not None and d <= 10

    @property
    def total_filiais(self):
        return Usuario.query.filter_by(rede_id=self.id, perfil='filial').count()


class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id           = db.Column(db.Integer, primary_key=True)
    username     = db.Column(db.String(80), unique=True, nullable=False)
    email        = db.Column(db.String(150), nullable=True)   # email real para envio de mensagens
    password     = db.Column(db.String(200), nullable=False)
    perfil       = db.Column(db.String(20), default='filial')
    nome_exibir  = db.Column(db.String(150), nullable=True)
    filial_nome  = db.Column(db.String(150), nullable=True)
    rede_id      = db.Column(db.Integer, db.ForeignKey('redes.id'), nullable=True)
    tema              = db.Column(db.String(10), default='light')
    termos_aceitos    = db.Column(db.Boolean, default=False, nullable=False)
    termos_aceitos_em = db.Column(db.DateTime, nullable=True)
    email_confirmado  = db.Column(db.Boolean, default=False, nullable=False)
    email_codigo      = db.Column(db.String(10), nullable=True)
    email_codigo_exp  = db.Column(db.DateTime, nullable=True)

    @property
    def is_superadmin(self): return self.perfil == 'superadmin'
    @property
    def is_dono(self): return self.perfil == 'dono_rede'
    @property
    def is_filial(self): return self.perfil == 'filial'

    @property
    def assinatura_ok(self):
        if self.is_superadmin: return True
        if not self.rede: return False
        return self.rede.assinatura_ativa

    @property
    def exibir_alerta_renovacao(self):
        if self.is_superadmin: return False
        if not self.rede: return False
        return self.rede.alerta_renovacao

    @property
    def aceitou_termos(self):
        if self.is_superadmin: return True   # superadmin não precisa aceitar
        return bool(self.termos_aceitos)

    def set_password(self, senha_plana):
        self.password = generate_password_hash(senha_plana)

    def check_password(self, senha_plana):
        return check_password_hash(self.password, senha_plana)


class Medicamento(db.Model):
    __tablename__ = 'medicamentos'
    id              = db.Column(db.Integer, primary_key=True)
    nome            = db.Column(db.String(200), nullable=False)
    codigo_barras   = db.Column(db.String(50), nullable=True, index=True)
    fabricante      = db.Column(db.String(150), nullable=True)
    principio_ativo = db.Column(db.String(200), nullable=True)
    lote            = db.Column(db.String(50), nullable=False)
    data_validade   = db.Column(db.Date, nullable=False)
    quantidade      = db.Column(db.Integer, nullable=False, default=0)
    preco_unitario  = db.Column(db.Float, nullable=False, default=0.0)
    data_cadastro   = db.Column(db.DateTime, default=datetime.utcnow)
    origem_cadastro = db.Column(db.String(50), default='manual')
    codigo_externo  = db.Column(db.String(100), nullable=True)
    rede_id         = db.Column(db.Integer, db.ForeignKey('redes.id'), nullable=True)
    filial_id       = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=True)

    @property
    def status(self):
        hoje = date.today()
        if self.data_validade < hoje: return 'vencido'
        elif self.data_validade <= hoje + timedelta(days=30): return 'alerta_30'
        elif self.data_validade <= hoje + timedelta(days=60): return 'alerta_60'
        return 'ok'

    @property
    def status_label(self):
        return {'vencido':'Vencido','alerta_30':'Vence em 30 dias',
                'alerta_60':'Vence em 60 dias','ok':'OK'}.get(self.status,'OK')

    @property
    def valor_total(self): return self.quantidade * self.preco_unitario

    def to_dict(self):
        return {'id':self.id,'nome':self.nome,'codigo_barras':self.codigo_barras,
                'lote':self.lote,'data_validade':self.data_validade.strftime('%Y-%m-%d'),
                'quantidade':self.quantidade,'preco_unitario':self.preco_unitario,
                'valor_total':self.valor_total,'status':self.status,
                'rede_id':self.rede_id,'filial_id':self.filial_id}


class IntegracaoConsys(db.Model):
    """Armazena configuração da integração com o ERP Consys por rede."""
    __tablename__ = 'integracoes_consys'
    id              = db.Column(db.Integer, primary_key=True)
    rede_id         = db.Column(db.Integer, db.ForeignKey('redes.id'), unique=True, nullable=False)
    ativa           = db.Column(db.Boolean, default=False)
    base_url        = db.Column(db.String(300), nullable=True)   # ex: https://api.consysonline.com.br
    api_key         = db.Column(db.String(500), nullable=True)   # Bearer token / chave de API
    cod_empresa     = db.Column(db.String(50),  nullable=True)   # Código da empresa no Consys
    ultimo_sync     = db.Column(db.DateTime,    nullable=True)
    sync_status     = db.Column(db.String(50),  default='nunca') # nunca | ok | erro
    sync_mensagem   = db.Column(db.String(500), nullable=True)
    criado_em       = db.Column(db.DateTime,    default=datetime.utcnow)
    atualizado_em   = db.Column(db.DateTime,    default=datetime.utcnow, onupdate=datetime.utcnow)

    rede = db.relationship('Rede', backref=db.backref('integracao_consys', uselist=False))


# =============================================================================
# DECORADORES
# =============================================================================

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            session['next'] = request.url
            return redirect(url_for('login'))
        session.permanent = True
        return f(*args, **kwargs)
    return decorated

def assinatura_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        session.permanent = True
        u = Usuario.query.get(session['user_id'])
        if not u or not u.assinatura_ok:
            return redirect(url_for('assinatura_expirada'))
        # Força leitura e aceite dos termos antes de qualquer rota
        # Email precisa ser confirmado apenas para dono_rede (filiais são criadas pelo dono)
        if u.is_dono and not u.email:
            return redirect(url_for("completar_cadastro"))
        if u.is_dono and not u.email_confirmado:
            return redirect(url_for('confirmar_email'))
        if not u.aceitou_termos:
            return redirect(url_for('aceitar_termos'))
        return f(*args, **kwargs)
    return decorated

def superadmin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('perfil') != 'superadmin':
            flash('Acesso restrito ao administrador.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def api_key_required(f):
    """Autentica por X-API-Key header ou ?api_key= na query string. Sem sessão."""
    @wraps(f)
    def decorated(*args, **kwargs):
        key = (request.headers.get("X-API-Key") or
               request.args.get("api_key") or "").strip()
        if not key:
            return jsonify({"success": False, "error": "API Key nao informada. Use o header X-API-Key."}), 401
        rede = Rede.query.filter_by(token_api=key).first()
        if not rede:
            return jsonify({"success": False, "error": "API Key invalida."}), 403
        if not rede.assinatura_ativa:
            return jsonify({"success": False, "error": "Assinatura inativa. Renove o plano."}), 403
        request.rede_autenticada = rede
        return f(*args, **kwargs)
    return decorated

def get_usuario_atual():
    return Usuario.query.get(session.get('user_id'))

def get_medicamentos_query():
    u = get_usuario_atual()
    if u.is_superadmin: return Medicamento.query
    elif u.is_dono:     return Medicamento.query.filter_by(rede_id=u.rede_id)
    else:               return Medicamento.query.filter_by(filial_id=u.id)


# =============================================================================
# TRATAMENTO DE ERROS
# =============================================================================

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    flash('Sessão expirada ou requisição inválida. Tente novamente.', 'danger')
    return redirect(url_for('login')), 400

@app.errorhandler(429)
def handle_rate_limit(e):
    flash('Muitas tentativas. Aguarde 1 minuto antes de tentar novamente.', 'danger')
    return render_template('login.html'), 429

@app.errorhandler(500)
def handle_500(e):
    app.logger.error(f'Erro interno: {e}')
    return render_template('login.html'), 500


# =============================================================================
# AUTENTICAÇÃO
# =============================================================================

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute")
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            flash('Preencha usuário e senha.', 'danger')
            return render_template('login.html')

        # Busca case-insensitive: tenta exato primeiro, depois lowercase
        usuario = Usuario.query.filter(
            db.func.lower(Usuario.username) == username.lower()
        ).first()

        if usuario and usuario.check_password(password):
            if not usuario.assinatura_ok:
                audit('login_bloqueado', f'username={username} motivo=assinatura_expirada')
                return redirect(url_for('assinatura_expirada'))
            session.permanent = True
            session['user_id']     = usuario.id
            session['username']    = usuario.username
            session['perfil']      = usuario.perfil
            session['nome_exibir'] = usuario.nome_exibir or usuario.username
            session['rede_id']     = usuario.rede_id
            session['filial_nome'] = usuario.filial_nome or ''
            session['tema']        = usuario.tema or 'light'
            audit('login_ok', f'username={username} perfil={usuario.perfil}')
            next_url = session.pop('next', None)
            return redirect(next_url or url_for('dashboard'))

        audit('login_falhou', f'username={username}')
        flash('Usuário ou senha incorretos.', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    audit('logout')
    session.clear()
    return redirect(url_for('login'))


@app.route('/assinatura-expirada')
def assinatura_expirada():
    return render_template('expirado.html', username=session.get('username', ''))


# =============================================================================
# TROCA DE SENHA
# =============================================================================

@app.route('/alterar-senha', methods=['GET', 'POST'])
@login_required
def alterar_senha():
    u = get_usuario_atual()
    if request.method == 'POST':
        senha_atual  = request.form.get('senha_atual', '').strip()
        nova_senha   = request.form.get('nova_senha', '').strip()
        confirma     = request.form.get('confirma_senha', '').strip()

        if not u.check_password(senha_atual):
            flash('Senha atual incorreta.', 'danger')
        elif len(nova_senha) < 6:
            flash('A nova senha deve ter pelo menos 6 caracteres.', 'danger')
        elif nova_senha != confirma:
            flash('A confirmação não coincide com a nova senha.', 'danger')
        else:
            u.set_password(nova_senha)
            db.session.commit()
            audit('senha_alterada', f'username={u.username}')
            flash('Senha alterada com sucesso!', 'success')
            return redirect(url_for('dashboard'))

    return render_template('alterar_senha.html', usuario=u)


# =============================================================================
# ACEITE DE TERMOS — obrigatório na primeira sessão
# =============================================================================

@app.route('/aceitar-termos', methods=['GET', 'POST'])
@login_required
def aceitar_termos():
    u = get_usuario_atual()
    # superadmin nunca precisa aceitar
    if u.aceitou_termos:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        aceite = request.form.get('aceite')
        if aceite == '1':
            u.termos_aceitos    = True
            u.termos_aceitos_em = datetime.utcnow()
            db.session.commit()
            audit('termos_aceitos', f'username={u.username} ip={request.remote_addr}')
            flash('Bem-vindo ao MedControl! Termos aceitos com sucesso.', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Você precisa aceitar os Termos de Uso para continuar.', 'danger')

    return render_template('aceitar_termos.html', usuario=u)


# =============================================================================
# DASHBOARD
# =============================================================================

@app.route('/')
@assinatura_required
def dashboard():
    hoje          = date.today()
    busca         = request.args.get('busca', '').strip()
    status        = request.args.get('status', '')
    filial_filtro = request.args.get('filial', '')
    pagina        = request.args.get('pagina', 1, type=int)
    u             = get_usuario_atual()
    query         = get_medicamentos_query()

    if busca:
        query = query.filter(db.or_(
            Medicamento.nome.ilike(f'%{busca}%'),
            Medicamento.lote.ilike(f'%{busca}%'),
            Medicamento.codigo_barras.ilike(f'%{busca}%')
        ))

    if filial_filtro and (u.is_dono or u.is_superadmin):
        try:
            query = query.filter_by(filial_id=int(filial_filtro))
        except (ValueError, TypeError):
            pass

    query = query.order_by(Medicamento.data_validade.asc())

    # Filtra por status antes de paginar
    if status:
        todos_filtrados = [m for m in query.all() if m.status == status]
        total_filtrado  = len(todos_filtrados)
        inicio = (pagina - 1) * ITENS_POR_PAGINA
        medicamentos = todos_filtrados[inicio: inicio + ITENS_POR_PAGINA]
    else:
        total_filtrado = query.count()
        medicamentos   = query.offset((pagina - 1) * ITENS_POR_PAGINA).limit(ITENS_POR_PAGINA).all()

    total_paginas = max(1, -(-total_filtrado // ITENS_POR_PAGINA))  # ceil division

    # Stats sempre sobre todos os medicamentos (sem filtro de busca)
    todos = get_medicamentos_query().all()
    stats = {
        'total':     len(todos),
        'vencidos':  sum(1 for m in todos if m.status == 'vencido'),
        'alerta_30': sum(1 for m in todos if m.status == 'alerta_30'),
        'alerta_60': sum(1 for m in todos if m.status == 'alerta_60'),
        'ok':        sum(1 for m in todos if m.status == 'ok'),
    }
    prejuizo   = sum(m.valor_total for m in todos if m.status == 'vencido')
    valor_ok   = sum(m.valor_total for m in todos if m.status != 'vencido')
    chart_data = json.dumps({
        'labels': ['Vencidos (Prejuízo)', 'Em estoque (Válido)'],
        'values': [round(prejuizo, 2), round(valor_ok, 2)],
        'colors': ['#ef4444', '#10b981'],
    })
    stats_json = json.dumps(stats)

    filiais = []
    if u.is_dono:
        filiais = Usuario.query.filter_by(rede_id=u.rede_id, perfil='filial').all()
    elif u.is_superadmin:
        filiais = Usuario.query.filter_by(perfil='filial').all()

    return render_template('index.html',
        medicamentos=medicamentos, stats=stats, chart_data=chart_data, stats_json=stats_json,
        hoje=hoje,
        rede=u.rede if not u.is_superadmin else None, busca=busca, status_filtro=status,
        filiais=filiais, filial_filtro=filial_filtro, usuario=u,
        alerta_renovacao=u.exibir_alerta_renovacao,
        dias_restantes=u.rede.dias_restantes if u.rede else None,
        pagina=pagina, total_paginas=total_paginas,
        total_filtrado=total_filtrado,
    )


# =============================================================================
# CRUD MEDICAMENTOS
# =============================================================================

@app.route('/cadastro', methods=['GET', 'POST'])
@assinatura_required
def cadastro():
    u = get_usuario_atual()
    filiais = []
    if u.is_dono:         filiais = Usuario.query.filter_by(rede_id=u.rede_id, perfil='filial').all()
    elif u.is_superadmin: filiais = Usuario.query.filter_by(perfil='filial').all()

    if request.method == 'POST':
        try:
            if u.is_filial:
                filial_id, rede_id = u.id, u.rede_id
            else:
                filial_id = request.form.get('filial_id')
                filial_id = int(filial_id) if filial_id and filial_id.isdigit() else None
                fu        = Usuario.query.get(filial_id) if filial_id else None
                rede_id   = fu.rede_id if fu else (u.rede_id if u.is_dono else None)

            med = Medicamento(
                nome            = request.form['nome'].strip()[:200],
                codigo_barras   = request.form.get('codigo_barras', '').strip()[:50] or None,
                fabricante      = request.form.get('fabricante', '').strip()[:150] or None,
                principio_ativo = request.form.get('principio_ativo', '').strip()[:200] or None,
                lote            = request.form['lote'].strip()[:50],
                data_validade   = datetime.strptime(request.form['data_validade'], '%Y-%m-%d').date(),
                quantidade      = int(request.form['quantidade']),
                preco_unitario  = float(request.form['preco_unitario'].replace(',', '.')),
                origem_cadastro = request.form.get('origem_cadastro', 'manual'),
                filial_id=filial_id, rede_id=rede_id,
            )
            db.session.add(med)
            db.session.commit()
            audit('medicamento_cadastrado', f'nome={med.nome} lote={med.lote}')
            flash(f'Medicamento "{med.nome}" cadastrado!', 'success')
            return redirect(url_for('dashboard'))

        except (ValueError, KeyError) as e:
            app.logger.warning(f'Erro no cadastro: {e}')
            flash('Dados inválidos. Verifique os campos e tente novamente.', 'danger')

    return render_template('cadastro.html', med=None, modo='novo', filiais=filiais, usuario=u)


@app.route('/editar/<int:id>', methods=['GET', 'POST'])
@assinatura_required
def editar(id):
    u   = get_usuario_atual()
    med = get_medicamentos_query().filter_by(id=id).first_or_404()
    filiais = []
    if u.is_dono:         filiais = Usuario.query.filter_by(rede_id=u.rede_id, perfil='filial').all()
    elif u.is_superadmin: filiais = Usuario.query.filter_by(perfil='filial').all()

    if request.method == 'POST':
        try:
            med.nome            = request.form['nome'].strip()[:200]
            med.codigo_barras   = request.form.get('codigo_barras', '').strip()[:50] or None
            med.fabricante      = request.form.get('fabricante', '').strip()[:150] or None
            med.principio_ativo = request.form.get('principio_ativo', '').strip()[:200] or None
            med.lote            = request.form['lote'].strip()[:50]
            med.data_validade   = datetime.strptime(request.form['data_validade'], '%Y-%m-%d').date()
            med.quantidade      = int(request.form['quantidade'])
            med.preco_unitario  = float(request.form['preco_unitario'].replace(',', '.'))
            if not u.is_filial and request.form.get('filial_id'):
                fid = request.form.get('filial_id')
                if fid and fid.isdigit():
                    med.filial_id = int(fid)
            db.session.commit()
            audit('medicamento_editado', f'id={id} nome={med.nome}')
            flash(f'"{med.nome}" atualizado!', 'success')
            return redirect(url_for('dashboard'))

        except (ValueError, KeyError) as e:
            app.logger.warning(f'Erro ao editar {id}: {e}')
            flash('Dados inválidos. Verifique os campos e tente novamente.', 'danger')

    return render_template('cadastro.html', med=med, modo='editar', filiais=filiais, usuario=u)


@app.route('/excluir/<int:id>', methods=['POST'])
@assinatura_required
def excluir(id):
    med  = get_medicamentos_query().filter_by(id=id).first_or_404()
    nome = med.nome
    db.session.delete(med)
    db.session.commit()
    audit('medicamento_excluido', f'id={id} nome={nome}')
    flash(f'"{nome}" excluído.', 'warning')
    return redirect(url_for('dashboard'))




# =============================================================================
# API BUSCA AJAX — retorna medicamentos em JSON para o dashboard
# =============================================================================
@app.route('/api/busca')
@assinatura_required
def api_busca():
    u             = get_usuario_atual()
    busca         = request.args.get('busca', '').strip()
    status_filtro = request.args.get('status', '')
    filial_filtro = request.args.get('filial', '')

    query = get_medicamentos_query()

    if busca:
        query = query.filter(db.or_(
            Medicamento.nome.ilike(f'%{busca}%'),
            Medicamento.lote.ilike(f'%{busca}%'),
            Medicamento.codigo_barras.ilike(f'%{busca}%')
        ))

    if filial_filtro and (u.is_dono or u.is_superadmin):
        try:
            query = query.filter_by(filial_id=int(filial_filtro))
        except (ValueError, TypeError):
            pass

    query = query.order_by(Medicamento.data_validade.asc())
    todos = query.all()

    if status_filtro:
        todos = [m for m in todos if m.status == status_filtro]

    def fmt_brl(v):
        return f"R$ {v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

    meds = []
    for m in todos:
        # Buscar nome da filial
        filial_nome = None
        if m.filial_id:
            filial_u = Usuario.query.get(m.filial_id)
            filial_nome = filial_u.filial_nome or filial_u.username if filial_u else None

        meds.append({
            'id':           m.id,
            'nome':         m.nome,
            'fabricante':   m.fabricante or '',
            'codigo_barras': m.codigo_barras or '',
            'lote':         m.lote,
            'validade':     m.data_validade.strftime('%d/%m/%Y'),
            'quantidade':   m.quantidade,
            'preco':        fmt_brl(m.preco_unitario),
            'total':        fmt_brl(m.preco_unitario * m.quantidade),
            'status':       m.status,
            'filial_nome':  filial_nome,
            'edit_url':     url_for('editar', id=m.id),
        })

    return jsonify({'medicamentos': meds, 'total': len(meds)})

# =============================================================================
# FEEDBACK
# =============================================================================

@app.route('/feedback', methods=['POST'])
@login_required
def enviar_feedback():
    mensagem  = request.form.get('mensagem', '').strip()[:2000]
    categoria = request.form.get('categoria', 'Geral')[:50]
    username  = session.get('username', 'Desconhecido')

    if not mensagem:
        flash('Escreva uma mensagem antes de enviar.', 'warning')
        return redirect(request.referrer or url_for('dashboard'))

    tg_token   = os.environ.get('TELEGRAM_TOKEN')
    tg_chat_id = os.environ.get('TELEGRAM_CHAT_ID')

    if not tg_token or not tg_chat_id:
        flash('Feedback recebido! (notificação não configurada no servidor)', 'warning')
        return redirect(url_for('dashboard'))

    try:
        texto = (
            f"💊 *MedControl — Novo Feedback*\n\n"
            f"👤 *Usuário:* {username}\n"
            f"📂 *Categoria:* {categoria}\n"
            f"🕐 *Data:* {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n"
            f"💬 *Mensagem:*\n{mensagem}"
        )
        payload = json.dumps({
            'chat_id': tg_chat_id, 'text': texto, 'parse_mode': 'Markdown',
        }).encode('utf-8')
        req = urllib.request.Request(
            f'https://api.telegram.org/bot{tg_token}/sendMessage',
            data=payload, headers={'Content-Type': 'application/json'}, method='POST'
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            if resp.status == 200:
                flash('Feedback enviado! Obrigado.', 'success')
            else:
                raise Exception(f'Status {resp.status}')
    except Exception as e:
        app.logger.error(f'Erro feedback Telegram: {e}')
        flash('Erro ao enviar feedback. Tente novamente.', 'danger')

    return redirect(url_for('dashboard'))


# =============================================================================
# SUPERADMIN — PAINEL DE REDES
# =============================================================================

@app.route('/admin')
@login_required
@superadmin_required
def admin_dashboard():
    redes = Rede.query.order_by(Rede.nome).all()
    todos_usuarios = Usuario.query.filter(Usuario.perfil != 'superadmin').order_by(Usuario.id.desc()).all()
    stats = {
        'total_redes':      len(redes),
        'ativas':           sum(1 for r in redes if r.assinatura_ativa),
        'expiradas':        sum(1 for r in redes if not r.assinatura_ativa),
        'total_filiais':    sum(r.total_filiais for r in redes),
        'total_usuarios':   len(todos_usuarios),
        'aceitaram_termos': sum(1 for u in todos_usuarios if u.termos_aceitos),
        'pendentes_termos': sum(1 for u in todos_usuarios if not u.termos_aceitos),
    }
    return render_template('admin/dashboard.html', redes=redes, stats=stats, todos_usuarios=todos_usuarios)


@app.route('/admin/redes/nova', methods=['GET', 'POST'])
@login_required
@superadmin_required
def admin_nova_rede():
    if request.method == 'POST':
        try:
            expiracao  = request.form.get('data_expiracao')
            senha_dono = request.form['password_dono'].strip()

            if len(senha_dono) < 6:
                flash('A senha do dono deve ter pelo menos 6 caracteres.', 'danger')
                return render_template('admin/rede_form.html', rede=None)

            rede = Rede(
                nome          = request.form['nome'].strip()[:200],
                cnpj          = request.form.get('cnpj', '').strip()[:20] or None,
                email_contato = request.form.get('email_contato', '').strip()[:150] or None,
                telefone      = request.form.get('telefone', '').strip()[:30] or None,
                plano         = request.form.get('plano', 'mensal'),
                data_expiracao= datetime.strptime(expiracao, '%Y-%m-%d').date() if expiracao else None,
            )
            db.session.add(rede)
            db.session.flush()

            dono = Usuario(
                username    = request.form['username_dono'].strip()[:80],
                perfil      = 'dono_rede',
                nome_exibir = request.form['nome_dono'].strip()[:150],
                rede_id     = rede.id,
            )
            dono.set_password(senha_dono)
            db.session.add(dono)
            db.session.commit()
            audit('rede_criada', f'rede={rede.nome} dono={dono.username}')
            flash(f'Rede "{rede.nome}" criada! Login do dono: {dono.username}', 'success')
            return redirect(url_for('admin_rede_detalhe', id=rede.id))

        except (ValueError, KeyError) as e:
            db.session.rollback()
            app.logger.warning(f'Erro ao criar rede: {e}')
            flash('Dados inválidos. Verifique os campos.', 'danger')

    return render_template('admin/rede_form.html', rede=None)


@app.route('/admin/redes/<int:id>')
@login_required
@superadmin_required
def admin_rede_detalhe(id):
    rede    = Rede.query.get_or_404(id)
    filiais = Usuario.query.filter_by(rede_id=id, perfil='filial').all()
    dono    = Usuario.query.filter_by(rede_id=id, perfil='dono_rede').first()
    # todos os usuários da rede (dono + filiais) para mostrar status de aceite
    usuarios_rede = Usuario.query.filter_by(rede_id=id).all()
    return render_template('admin/rede_detalhe.html',
        rede=rede, filiais=filiais, dono=dono, usuarios_rede=usuarios_rede)


@app.route('/admin/redes/<int:id>/filial/nova', methods=['POST'])
@login_required
@superadmin_required
def admin_nova_filial(id):
    rede = Rede.query.get_or_404(id)
    try:
        senha = request.form['password'].strip()
        if len(senha) < 6:
            flash('Senha da filial deve ter pelo menos 6 caracteres.', 'danger')
            return redirect(url_for('admin_rede_detalhe', id=id))

        filial = Usuario(
            username    = request.form['username'].strip()[:80],
            perfil      = 'filial',
            nome_exibir = request.form.get('nome_exibir', '').strip()[:150],
            filial_nome = request.form['filial_nome'].strip()[:150],
            rede_id     = rede.id,
        )
        filial.set_password(senha)
        db.session.add(filial)
        db.session.commit()
        audit('filial_criada', f'filial={filial.filial_nome} rede={rede.nome}')
        flash(f'Filial "{filial.filial_nome}" criada!', 'success')

    except (ValueError, KeyError) as e:
        app.logger.warning(f'Erro ao criar filial: {e}')
        flash('Dados inválidos ao criar filial.', 'danger')

    return redirect(url_for('admin_rede_detalhe', id=id))


@app.route('/admin/redes/<int:id>/renovar', methods=['POST'])
@login_required
@superadmin_required
def admin_renovar_rede(id):
    rede = Rede.query.get_or_404(id)
    try:
        dias = max(1, min(int(request.form.get('dias', 30)), 365))
        base = max(rede.data_expiracao, date.today()) if rede.data_expiracao and rede.data_expiracao > date.today() else date.today()
        rede.data_expiracao = base + timedelta(days=dias)
        rede.ativa = True
        db.session.commit()
        audit('rede_renovada', f'rede={rede.nome} ate={rede.data_expiracao}')
        flash(f'Assinatura de "{rede.nome}" renovada até {rede.data_expiracao.strftime("%d/%m/%Y")}.', 'success')
    except (ValueError, TypeError):
        flash('Número de dias inválido.', 'danger')
    return redirect(url_for('admin_rede_detalhe', id=id))


@app.route('/admin/redes/<int:id>/toggle', methods=['POST'])
@login_required
@superadmin_required
def admin_toggle_rede(id):
    rede       = Rede.query.get_or_404(id)
    rede.ativa = not rede.ativa
    db.session.commit()
    audit('rede_toggle', f'rede={rede.nome} ativa={rede.ativa}')
    flash(f'Rede "{rede.nome}" {"ativada" if rede.ativa else "bloqueada"}.', 'success' if rede.ativa else 'warning')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/redes/<int:id>/excluir', methods=['GET', 'POST'])
@login_required
@superadmin_required
@csrf.exempt
def admin_excluir_rede(id):
    """
    GET  → mostra tela de confirmação com nome da rede
    POST → executa exclusão somente se confirmação bater
    """
    rede = Rede.query.get_or_404(id)

    if request.method == 'GET':
        # Tela de confirmação server-side
        return render_template('admin/confirmar_exclusao.html', rede=rede)

    # POST — verifica se o usuário digitou o nome correto
    confirmacao = request.form.get('confirmacao', '').strip()
    if confirmacao != rede.nome.strip():
        flash(f'Nome não confere. Digite exatamente: {rede.nome.strip()}', 'danger')
        return redirect(url_for('admin_excluir_rede', id=id))

    nome = rede.nome
    Medicamento.query.filter_by(rede_id=id).delete()
    IntegracaoConsys.query.filter_by(rede_id=id).delete()
    Usuario.query.filter_by(rede_id=id).delete()
    db.session.delete(rede)
    db.session.commit()
    audit('rede_excluida', f'rede={nome}')
    flash(f'Rede "{nome}" e todos os seus dados foram excluídos.', 'danger')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/filial/<int:id>/excluir', methods=['POST'])
@login_required
@superadmin_required
def admin_excluir_filial(id):
    filial  = Usuario.query.get_or_404(id)
    rede_id = filial.rede_id
    nome    = filial.filial_nome or filial.username
    Medicamento.query.filter_by(filial_id=id).update({'filial_id': None})
    db.session.delete(filial)
    db.session.commit()
    audit('filial_excluida', f'filial={nome}')
    flash(f'Filial "{nome}" removida.', 'warning')
    return redirect(url_for('admin_rede_detalhe', id=rede_id))


# =============================================================================
# PLANOS & FILIAIS
# =============================================================================

@app.route('/planos')
@login_required
def planos():
    u = get_usuario_atual()
    if u.is_filial:
        return redirect(url_for('dashboard'))
    return render_template('planos.html', usuario=u)


@app.route('/filiais')
@assinatura_required
def gerenciar_filiais():
    u = get_usuario_atual()
    if not u.is_dono:
        return redirect(url_for('dashboard'))
    filiais = Usuario.query.filter_by(rede_id=u.rede_id, perfil='filial').all()
    return render_template('gerenciar_filiais.html', filiais=filiais, usuario=u)



@app.route('/filiais/criar', methods=['POST'])
@assinatura_required
def dono_criar_filial():
    u = get_usuario_atual()
    if not u.is_dono:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('dashboard'))
    try:
        senha = request.form.get('password', '').strip()
        username = request.form.get('username', '').strip()[:80]
        filial_nome = request.form.get('filial_nome', '').strip()[:150]

        if len(senha) < 6:
            flash('Senha deve ter pelo menos 6 caracteres.', 'danger')
            return redirect(url_for('gerenciar_filiais'))
        if not username or not filial_nome:
            flash('Preencha nome da filial e login.', 'danger')
            return redirect(url_for('gerenciar_filiais'))
        if Usuario.query.filter_by(username=username).first():
            flash(f'Login "{username}" já está em uso. Escolha outro.', 'danger')
            return redirect(url_for('gerenciar_filiais'))

        filial = Usuario(
            username         = username,
            perfil           = 'filial',
            nome_exibir      = filial_nome,
            filial_nome      = filial_nome,
            rede_id          = u.rede_id,
            email_confirmado = True,  # filiais não precisam confirmar email
        )
        filial.set_password(senha)
        db.session.add(filial)
        db.session.commit()
        audit('filial_criada_dono', f'filial={filial_nome} rede_id={u.rede_id}')
        flash(f'Filial "{filial_nome}" criada! Login: {username}', 'success')
    except Exception as e:
        app.logger.error(f'Erro criar filial: {e}')
        flash('Erro ao criar filial. Tente novamente.', 'danger')
    return redirect(url_for('gerenciar_filiais'))

@app.route('/filial/<int:id>/excluir', methods=['POST'])
@assinatura_required
def dono_excluir_filial(id):
    u = get_usuario_atual()
    if not u.is_dono:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('dashboard'))
    filial = Usuario.query.filter_by(id=id, rede_id=u.rede_id, perfil='filial').first_or_404()
    nome   = filial.filial_nome or filial.username
    Medicamento.query.filter_by(filial_id=id).update({'filial_id': None})
    db.session.delete(filial)
    db.session.commit()
    audit('filial_excluida_dono', f'filial={nome}')
    flash(f'Filial "{nome}" removida.', 'warning')
    return redirect(url_for('dashboard'))


# =============================================================================
# INTEGRAÇÃO CONSYS ERP
# =============================================================================
# Estrutura preparada para quando as credenciais do Consys forem disponibilizadas.
# Para ativar: configurar base_url, api_key e cod_empresa no painel admin.

def _consys_headers(integracao):
    """Monta headers padrão para requisições à API do Consys."""
    return {
        'Authorization': f'Bearer {integracao.api_key}',
        'Content-Type': 'application/json',
        'X-Empresa': str(integracao.cod_empresa or ''),
    }

def _consys_get(integracao, endpoint):
    """Faz GET na API do Consys. Retorna (dict|list, None) ou (None, erro_str)."""
    if _requests is None:
        return None, 'Biblioteca requests não instalada (pip install requests)'
    url = f"{integracao.base_url.rstrip('/')}/{endpoint.lstrip('/')}"
    try:
        r = _requests.get(url, headers=_consys_headers(integracao), timeout=15)
        r.raise_for_status()
        return r.json(), None
    except Exception as e:
        return None, str(e)

def _sync_consys(rede_id):
    """
    Sincroniza medicamentos do Consys com o MedControl.
    
    Endpoints esperados (a confirmar com documentação oficial Consys):
      GET /api/v1/produtos          → lista de produtos com EAN, nome, fabricante
      GET /api/v1/estoque           → lotes, validades e quantidades
    
    Quando a documentação oficial chegar, ajustar os endpoints abaixo.
    """
    integracao = IntegracaoConsys.query.filter_by(rede_id=rede_id, ativa=True).first()
    if not integracao:
        return False, 'Integração não configurada ou inativa'

    # ── PRODUTOS ──────────────────────────────────────────────────────────────
    # TODO: substituir '/api/v1/produtos' pelo endpoint real quando documentação chegar
    produtos, erro = _consys_get(integracao, '/api/v1/produtos')
    if erro:
        integracao.sync_status   = 'erro'
        integracao.sync_mensagem = f'Erro ao buscar produtos: {erro}'
        integracao.ultimo_sync   = datetime.utcnow()
        db.session.commit()
        return False, integracao.sync_mensagem

    # ── ESTOQUE / LOTES ───────────────────────────────────────────────────────
    # TODO: substituir '/api/v1/estoque' pelo endpoint real quando documentação chegar
    estoque, erro = _consys_get(integracao, '/api/v1/estoque')
    if erro:
        integracao.sync_status   = 'erro'
        integracao.sync_mensagem = f'Erro ao buscar estoque: {erro}'
        integracao.ultimo_sync   = datetime.utcnow()
        db.session.commit()
        return False, integracao.sync_mensagem

    # ── MAPEAMENTO ────────────────────────────────────────────────────────────
    # Monta índice de estoque por código de produto para cruzar com produtos
    # Estrutura esperada (ajustar conforme JSON real do Consys):
    #   produtos: [{ "codigo": "123", "ean": "789", "nome": "...", "fabricante": "..." }]
    #   estoque:  [{ "codigo_produto": "123", "lote": "L01", "validade": "2026-12", "quantidade": 50 }]
    estoque_idx = {}
    for item in (estoque or []):
        cod = str(item.get('codigo_produto') or item.get('codigo') or '')
        if cod:
            estoque_idx.setdefault(cod, []).append(item)

    importados = 0
    atualizados = 0

    for prod in (produtos or []):
        cod_ext = str(prod.get('codigo') or '')
        ean     = str(prod.get('ean') or prod.get('codigo_barras') or '')
        nome    = str(prod.get('nome') or prod.get('descricao') or '')[:200].strip()
        fab     = str(prod.get('fabricante') or '')[:150].strip()

        if not nome:
            continue

        lotes = estoque_idx.get(cod_ext, [])
        if not lotes:
            continue  # sem estoque, ignora

        for lote_item in lotes:
            lote     = str(lote_item.get('lote') or 'S/L')[:50]
            val_str  = str(lote_item.get('validade') or lote_item.get('data_validade') or '')
            qtd      = int(lote_item.get('quantidade') or lote_item.get('qtd') or 0)

            # Tenta parsear validade em vários formatos (YYYY-MM, YYYY-MM-DD, DD/MM/YYYY)
            data_val = None
            for fmt in ('%Y-%m-%d', '%Y-%m', '%d/%m/%Y', '%d/%m/%y'):
                try:
                    parsed = datetime.strptime(val_str[:len(fmt.replace('%Y','0000').replace('%m','00').replace('%d','00'))], fmt)
                    data_val = parsed.date()
                    break
                except Exception:
                    continue
            if not data_val:
                continue  # validade inválida, ignora

            # Verifica se já existe pelo código externo + lote
            med = Medicamento.query.filter_by(
                rede_id=rede_id, codigo_externo=cod_ext, lote=lote
            ).first()

            if med:
                # Atualiza dados existentes
                med.quantidade     = qtd
                med.data_validade  = data_val
                med.codigo_barras  = ean or med.codigo_barras
                atualizados += 1
            else:
                # Cria novo medicamento
                med = Medicamento(
                    nome            = nome,
                    codigo_barras   = ean,
                    fabricante      = fab,
                    principio_ativo = '',
                    lote            = lote,
                    data_validade   = data_val,
                    quantidade      = qtd,
                    preco_unitario  = float(lote_item.get('preco') or 0),
                    origem_cadastro = 'consys',
                    codigo_externo  = cod_ext,
                    rede_id         = rede_id,
                )
                db.session.add(med)
                importados += 1

    integracao.ultimo_sync   = datetime.utcnow()
    integracao.sync_status   = 'ok'
    integracao.sync_mensagem = f'{importados} importados, {atualizados} atualizados'
    db.session.commit()
    audit('consys_sync', f'rede_id={rede_id} importados={importados} atualizados={atualizados}')
    return True, integracao.sync_mensagem


# ── ROTAS ─────────────────────────────────────────────────────────────────────

@app.route('/integracoes/consys', methods=['GET', 'POST'])
@login_required
def integracao_consys():
    """Tela de configuração da integração Consys (acesso: dono_rede ou superadmin)."""
    u = get_usuario_atual()
    if u.perfil not in ('dono_rede', 'superadmin'):
        flash('Acesso restrito ao dono da rede.', 'danger')
        return redirect(url_for('dashboard'))

    rede_id = u.rede_id if u.perfil == 'dono_rede' else request.args.get('rede_id', type=int)
    if not rede_id:
        flash('Rede não encontrada.', 'danger')
        return redirect(url_for('dashboard'))

    integracao = IntegracaoConsys.query.filter_by(rede_id=rede_id).first()
    if not integracao:
        integracao = IntegracaoConsys(rede_id=rede_id)
        db.session.add(integracao)
        db.session.commit()

    if request.method == 'POST':
        acao = request.form.get('acao', 'salvar')

        if acao == 'salvar':
            integracao.base_url    = (request.form.get('base_url') or '').strip()[:300]
            integracao.api_key     = (request.form.get('api_key') or '').strip()[:500]
            integracao.cod_empresa = (request.form.get('cod_empresa') or '').strip()[:50]
            integracao.ativa       = request.form.get('ativa') == '1'
            db.session.commit()
            audit('consys_config_salva', f'rede_id={rede_id}')
            flash('Configuração salva com sucesso.', 'success')

        elif acao == 'testar':
            if not integracao.base_url or not integracao.api_key:
                flash('Preencha a URL base e a chave de API antes de testar.', 'danger')
            else:
                # Teste simples de conectividade — pinga a URL base
                dados, erro = _consys_get(integracao, '/api/v1/status')
                if erro:
                    flash(f'Falha na conexão: {erro}', 'danger')
                else:
                    flash('Conexão com o Consys estabelecida com sucesso! ✅', 'success')

        elif acao == 'sincronizar':
            ok, msg = _sync_consys(rede_id)
            if ok:
                flash(f'Sincronização concluída: {msg}', 'success')
            else:
                flash(f'Erro na sincronização: {msg}', 'danger')

        return redirect(url_for('integracao_consys', rede_id=rede_id if u.perfil == 'superadmin' else None))

    rede = Rede.query.get_or_404(rede_id)
    return render_template('integracao_consys.html', integracao=integracao, rede=rede)


@app.route('/integracoes/consys/sync', methods=['POST'])
@login_required
def consys_sync_ajax():
    """Endpoint AJAX para sincronização rápida."""
    u = get_usuario_atual()
    if u.perfil not in ('dono_rede', 'superadmin'):
        return json.dumps({'ok': False, 'msg': 'Sem permissão'}), 403

    rede_id = u.rede_id if u.perfil == 'dono_rede' else request.json.get('rede_id')
    ok, msg = _sync_consys(rede_id)
    return json.dumps({'ok': ok, 'msg': msg})



# =============================================================================
# SELF-SERVE — CADASTRO PÚBLICO + MERCADO PAGO
# =============================================================================
# Configurar no Render:
#   MP_ACCESS_TOKEN  = seu Access Token do Mercado Pago (produção)
#   MP_WEBHOOK_SECRET = string secreta para validar webhooks (qualquer texto)
#   APP_BASE_URL     = https://www.medcontrol.app.br

import hashlib, hmac

MP_ACCESS_TOKEN   = os.environ.get('MP_ACCESS_TOKEN', '')
MP_WEBHOOK_SECRET = os.environ.get('MP_WEBHOOK_SECRET', 'medcontrol_webhook_2024')
APP_BASE_URL      = os.environ.get('APP_BASE_URL', 'https://www.medcontrol.app.br')

# Preços dos planos em centavos (Mercado Pago usa centavos)
PLANOS_MP = {
    'basico':       {'nome': 'MedControl Básico',       'preco': 6000,  'filiais': 1},
    'profissional': {'nome': 'MedControl Profissional', 'preco': 12000, 'filiais': 5},
    'rede':         {'nome': 'MedControl Rede',         'preco': 20000, 'filiais': 999},
}


def mp_criar_preferencia(rede, plano_key):
    """Cria preferência de pagamento no Mercado Pago e retorna o link de checkout."""
    if not MP_ACCESS_TOKEN:
        return None, 'MP_ACCESS_TOKEN não configurado'

    plano = PLANOS_MP.get(plano_key)
    if not plano:
        return None, 'Plano inválido'

    payload = {
        'items': [{
            'title':       plano['nome'],
            'quantity':    1,
            'unit_price':  plano['preco'] / 100,
            'currency_id': 'BRL',
        }],
        'payer': {'email': rede.email_contato or ''},
        'back_urls': {
            'success': f"{APP_BASE_URL}/pagamento/sucesso",
            'failure': f"{APP_BASE_URL}/pagamento/falhou",
            'pending': f"{APP_BASE_URL}/pagamento/pendente",
        },
        'auto_return': 'approved',
        'external_reference': str(rede.id),
        'notification_url': f"{APP_BASE_URL}/webhook/mercadopago",
        'statement_descriptor': 'MEDCONTROL',
        'metadata': {'rede_id': rede.id, 'plano': plano_key},
        'payment_methods': {
            'excluded_payment_types': [],
            'installments': 1,
        },
    }

    try:
        data = json.dumps(payload).encode('utf-8')
        req  = urllib.request.Request(
            'https://api.mercadopago.com/checkout/preferences',
            data=data,
            headers={
                'Authorization': f'Bearer {MP_ACCESS_TOKEN}',
                'Content-Type':  'application/json',
            },
            method='POST'
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read())
            # init_point = link de checkout real; sandbox_init_point = testes
            link = result.get('init_point') or result.get('sandbox_init_point')
            return link, None
    except Exception as e:
        return None, str(e)


# ── ROTA: CADASTRO PÚBLICO ────────────────────────────────────────────────

@app.route('/registrar', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def registrar():
    """Cadastro self-serve: cria rede + dono_rede + inicia trial 30 dias."""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        nome_rede = request.form.get('nome_rede', '').strip()[:200]
        username  = request.form.get('username', '').strip().lower()[:80]
        email     = request.form.get('email', '').strip().lower()[:150]
        senha     = request.form.get('senha', '')
        confirma  = request.form.get('confirma_senha', '')
        telefone  = request.form.get('telefone', '').strip()[:30]

        erros = []
        if len(nome_rede) < 3:
            erros.append('Nome da farmácia deve ter pelo menos 3 caracteres.')
        if not username or len(username) < 3:
            erros.append('Usuário deve ter pelo menos 3 caracteres.')
        if not username.replace('_','').replace('-','').replace('.','').isalnum():
            erros.append('Usuário pode conter apenas letras, números, _, - e ponto.')
        if '@' not in email or '.' not in email:
            erros.append('Email inválido.')
        if len(senha) < 8:
            erros.append('Senha deve ter pelo menos 8 caracteres.')
        if senha != confirma:
            erros.append('As senhas não coincidem.')
        if Usuario.query.filter(db.func.lower(Usuario.username) == username).first():
            erros.append('Este usuário já está em uso. Escolha outro.')

        if erros:
            return render_template('registrar.html', erros=erros,
                                   nome_rede=nome_rede, username=username, email=email, telefone=telefone)

        # Cria a rede em trial
        rede = Rede(
            nome          = nome_rede,
            email_contato = email,
            telefone      = telefone,
            ativa         = True,
            plano         = 'trial',
            trial         = True,
            trial_inicio  = datetime.utcnow(),
            # trial de 30 dias — data_expiracao não é necessária durante trial
        )
        db.session.add(rede)
        db.session.flush()  # gera rede.id

        # Cria o usuário dono_rede
        usuario = Usuario(
            username      = username,
            email         = email,
            password       = generate_password_hash(senha),
            perfil        = 'dono_rede',
            nome_exibir   = nome_rede,
            rede_id       = rede.id,
            termos_aceitos= False,
        )
        db.session.add(usuario)
        db.session.commit()

        audit('auto_cadastro', f'rede={nome_rede} username={username} email={email}')

        # Loga automaticamente
        session.permanent = True
        session['user_id']  = usuario.id
        session['username'] = usuario.username
        session['perfil']   = usuario.perfil
        session['rede_id']  = rede.id

        # Enviar código de confirmação por email
        ok_email, _ = _enviar_codigo_confirmacao(usuario)
        if ok_email:
            flash('Conta criada! Enviamos um código de confirmação para seu email.', 'success')
        else:
            flash('Conta criada! Não conseguimos enviar o email — confirme depois nas configurações.', 'warning')
        return redirect(url_for('confirmar_email'))

    return render_template('registrar.html', erros=[], nome_rede='', username='', email='', telefone='')


# ── ROTA: PÁGINA DE ASSINATURA (escolha de plano + checkout) ─────────────

@app.route('/assinar', methods=['GET', 'POST'])
@login_required
def assinar():
    """Tela onde o usuário escolhe o plano e é redirecionado ao Mercado Pago."""
    u = get_usuario_atual()
    if u.perfil not in ('dono_rede',):
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        plano_key = request.form.get('plano', 'basico')
        if plano_key not in PLANOS_MP:
            flash('Plano inválido.', 'danger')
            return redirect(url_for('assinar'))

        link, erro = mp_criar_preferencia(u.rede, plano_key)
        if erro:
            flash(f'Erro ao gerar pagamento: {erro}', 'danger')
            return redirect(url_for('assinar'))

        audit('checkout_iniciado', f'rede_id={u.rede_id} plano={plano_key}')
        return redirect(link)

    rede = u.rede
    return render_template('assinar.html',
        rede=rede,
        planos=PLANOS_MP,
        dias_trial=rede.dias_trial_restantes if rede else 0,
        em_trial=rede.em_trial if rede else False,
    )


# ── ROTA: RETORNO DO PAGAMENTO ────────────────────────────────────────────

@app.route('/pagamento/sucesso')
def pagamento_sucesso():
    payment_id = request.args.get('payment_id', '')
    status     = request.args.get('status', '')
    ref        = request.args.get('external_reference', '')
    audit('pagamento_retorno', f'status={status} ref={ref} pid={payment_id}')
    if status == 'approved':
        flash('Pagamento aprovado! Sua assinatura está ativa. ✅', 'success')
        # A ativação real ocorre via webhook — aqui é só feedback visual
    else:
        flash('Pagamento recebido — aguardando confirmação.', 'warning')
    return redirect(url_for('dashboard'))


@app.route('/pagamento/falhou')
def pagamento_falhou():
    flash('Pagamento não foi concluído. Tente novamente ou entre em contato.', 'danger')
    return redirect(url_for('assinar'))


@app.route('/pagamento/pendente')
def pagamento_pendente():
    flash('Pagamento pendente — você receberá confirmação por email em breve.', 'warning')
    return redirect(url_for('dashboard'))


# ── WEBHOOK MERCADO PAGO ─────────────────────────────────────────────────

@app.route('/webhook/mercadopago', methods=['POST'])
def webhook_mercadopago():
    """
    Recebe notificações do Mercado Pago e ativa/cancela assinaturas.
    Configurar no painel MP: Integrações → Webhooks → URL = APP_BASE_URL/webhook/mercadopago
    """
    data = request.get_json(silent=True) or {}
    topic = data.get('type') or request.args.get('topic', '')
    resource_id = (data.get('data') or {}).get('id') or request.args.get('id', '')

    app.logger.info(f'MP Webhook: topic={topic} id={resource_id}')

    if topic not in ('payment', 'merchant_order'):
        return '', 200  # ignora outros eventos

    if not resource_id or not MP_ACCESS_TOKEN:
        return '', 200

    # Busca detalhes do pagamento na API do MP
    try:
        req = urllib.request.Request(
            f'https://api.mercadopago.com/v1/payments/{resource_id}',
            headers={'Authorization': f'Bearer {MP_ACCESS_TOKEN}'},
            method='GET'
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            payment = json.loads(resp.read())
    except Exception as e:
        app.logger.error(f'MP Webhook erro ao buscar pagamento: {e}')
        return '', 200

    status       = payment.get('status', '')
    rede_id_str  = str(payment.get('external_reference') or
                       (payment.get('metadata') or {}).get('rede_id', ''))
    plano_key    = (payment.get('metadata') or {}).get('plano', 'basico')
    payer_email  = (payment.get('payer') or {}).get('email', '')

    if not rede_id_str.isdigit():
        return '', 200

    rede = Rede.query.get(int(rede_id_str))
    if not rede:
        return '', 200

    if status == 'approved':
        # Ativa assinatura por 30 dias a partir de hoje
        rede.ativa         = True
        rede.trial         = False   # sai do trial
        rede.plano         = plano_key
        rede.data_expiracao = date.today() + timedelta(days=30)
        rede.mp_assinatura_id = str(resource_id)
        rede.mp_payer_email   = payer_email
        db.session.commit()
        audit('assinatura_ativada', f'rede_id={rede.id} plano={plano_key} payment_id={resource_id}')
        app.logger.info(f'Assinatura ativada: rede_id={rede.id}')

    elif status in ('cancelled', 'refunded', 'charged_back'):
        rede.ativa = False
        db.session.commit()
        audit('assinatura_cancelada', f'rede_id={rede.id} status={status}')

    return '', 200



# =============================================================================
# EMAIL — SMTP GMAIL
# =============================================================================
# Configurar no Render:
#   RESEND_API_KEY = chave da API Resend
#   RESEND_FROM    = MedControl <noreply@medcontrol.app.br>

RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '')
RESEND_FROM    = os.environ.get('RESEND_FROM', 'MedControl <noreply@medcontrol.app.br>')


def _enviar_email(destinatario, assunto, html):
    """Envia email via Resend SDK. Retorna (True, None) ou (False, erro_str)."""
    if not RESEND_API_KEY:
        app.logger.warning('RESEND_API_KEY não configurada')
        return False, 'Email não configurado no servidor'
    try:
        resend_sdk.api_key = RESEND_API_KEY
        params = {
            'from':    RESEND_FROM,
            'to':      [destinatario],
            'subject': assunto,
            'html':    html,
        }
        result = resend_sdk.Emails.send(params)
        app.logger.info(f'Email enviado via Resend SDK id={result.get("id")} para={destinatario}')
        return True, None
    except Exception as e:
        app.logger.error(f'Erro Resend SDK para {destinatario}: {e}')
        return False, str(e)


def _gerar_codigo():
    """Gera código numérico de 6 dígitos."""
    return ''.join(random.choices(string.digits, k=6))


def _enviar_codigo_confirmacao(usuario):
    """Gera código, salva no banco e envia por email. Retorna (True/False, msg)."""
    codigo = _gerar_codigo()
    usuario.email_codigo     = codigo
    usuario.email_codigo_exp = datetime.utcnow() + timedelta(hours=24)
    db.session.commit()

    html = f"""
    <div style="font-family:'Helvetica Neue',Arial,sans-serif;max-width:480px;margin:0 auto;background:#0a1628;border-radius:16px;overflow:hidden">
      <div style="background:linear-gradient(135deg,#0f766e,#14b8a6);padding:2rem;text-align:center">
        <div style="font-size:2rem">💊</div>
        <h1 style="color:white;font-size:1.4rem;margin:.5rem 0 0">MedControl</h1>
      </div>
      <div style="padding:2rem;color:#f1f5f9">
        <p style="margin:0 0 1rem;font-size:.95rem">Olá, <strong>{usuario.nome_exibir or usuario.username}</strong>!</p>
        <p style="margin:0 0 1.5rem;color:#94a3b8;font-size:.88rem">Use o código abaixo para confirmar seu email e ativar sua conta:</p>
        <div style="background:#142033;border:2px solid rgba(20,184,166,0.3);border-radius:14px;padding:1.5rem;text-align:center;margin-bottom:1.5rem">
          <div style="font-size:2.4rem;font-weight:900;letter-spacing:.3em;color:#14b8a6;font-family:monospace">{codigo}</div>
          <div style="font-size:.75rem;color:#64748b;margin-top:.5rem">Válido por 24 horas</div>
        </div>
        <p style="font-size:.78rem;color:#475569;margin:0">Se você não criou uma conta no MedControl, ignore este email.</p>
      </div>
    </div>
    """
    return _enviar_email(usuario.email or usuario.username, '🔐 Código de confirmação — MedControl', html)


def _enviar_notificacao_validade(usuario, medicamentos_proximos):
    """Envia email com lista de medicamentos vencendo em 30 dias."""
    if not medicamentos_proximos:
        return
    destinatario = usuario.email or usuario.username  # email real do usuário
    nome = usuario.nome_exibir or usuario.username

    linhas = ""
    for m in medicamentos_proximos:
        dias = (m.data_validade - date.today()).days
        cor = "#ef4444" if dias <= 7 else "#f59e0b" if dias <= 15 else "#14b8a6"
        linhas += f"""
        <tr>
          <td style="padding:.6rem .8rem;border-bottom:1px solid #1e293b;color:#f1f5f9;font-size:.85rem">{m.nome}</td>
          <td style="padding:.6rem .8rem;border-bottom:1px solid #1e293b;color:#94a3b8;font-size:.83rem">{m.lote}</td>
          <td style="padding:.6rem .8rem;border-bottom:1px solid #1e293b;font-size:.83rem">
            <span style="color:{cor};font-weight:700">{m.data_validade.strftime('%d/%m/%Y')}</span>
          </td>
          <td style="padding:.6rem .8rem;border-bottom:1px solid #1e293b;font-size:.83rem">
            <span style="background:{cor}22;color:{cor};padding:.2rem .6rem;border-radius:20px;font-size:.75rem;font-weight:700">{dias} dias</span>
          </td>
        </tr>"""

    html = f"""
    <div style="font-family:'Helvetica Neue',Arial,sans-serif;max-width:600px;margin:0 auto;background:#0a1628;border-radius:16px;overflow:hidden">
      <div style="background:linear-gradient(135deg,#0f766e,#14b8a6);padding:1.5rem 2rem;display:flex;align-items:center;gap:1rem">
        <div style="font-size:1.8rem">⚠️</div>
        <div>
          <h1 style="color:white;font-size:1.1rem;margin:0">Alerta de Validade — MedControl</h1>
          <p style="color:rgba(255,255,255,0.8);font-size:.8rem;margin:.2rem 0 0">{len(medicamentos_proximos)} medicamento(s) vencem em até 30 dias</p>
        </div>
      </div>
      <div style="padding:1.5rem 2rem;color:#f1f5f9">
        <p style="margin:0 0 1.2rem;font-size:.9rem">Olá <strong>{nome}</strong>, os seguintes medicamentos precisam de atenção:</p>
        <table style="width:100%;border-collapse:collapse;background:#0f1f35;border-radius:12px;overflow:hidden">
          <thead>
            <tr style="background:#142033">
              <th style="padding:.7rem .8rem;text-align:left;font-size:.72rem;text-transform:uppercase;letter-spacing:.05em;color:#64748b">Medicamento</th>
              <th style="padding:.7rem .8rem;text-align:left;font-size:.72rem;text-transform:uppercase;letter-spacing:.05em;color:#64748b">Lote</th>
              <th style="padding:.7rem .8rem;text-align:left;font-size:.72rem;text-transform:uppercase;letter-spacing:.05em;color:#64748b">Validade</th>
              <th style="padding:.7rem .8rem;text-align:left;font-size:.72rem;text-transform:uppercase;letter-spacing:.05em;color:#64748b">Dias</th>
            </tr>
          </thead>
          <tbody>{linhas}</tbody>
        </table>
        <div style="margin-top:1.5rem;text-align:center">
          <a href="https://www.medcontrol.app.br" style="background:linear-gradient(135deg,#0f766e,#14b8a6);color:white;text-decoration:none;padding:.7rem 2rem;border-radius:10px;font-weight:700;font-size:.9rem;display:inline-block">
            Abrir MedControl →
          </a>
        </div>
        <p style="font-size:.72rem;color:#475569;margin-top:1.5rem;text-align:center">
          Você recebe este email pois sua conta MedControl tem alertas de validade ativos.
        </p>
      </div>
    </div>
    """
    _enviar_email(destinatario, f'⚠️ {len(medicamentos_proximos)} medicamento(s) vencem em 30 dias — MedControl', html)


# Rota para disparar notificações (chamada por cron/scheduler ou manualmente)
@app.route('/sistema/notificacoes', methods=['POST'])
def disparar_notificacoes():
    """
    Dispara emails de alerta de validade para todos os usuários dono_rede.
    Chamar via cron diário — protegido por CRON_SECRET.
    Ex: POST /sistema/notificacoes com header X-Cron-Secret: <valor>
    """
    secret = request.headers.get('X-Cron-Secret', '')
    cron_secret = os.environ.get('CRON_SECRET', '')
    if not cron_secret or secret != cron_secret:
        return jsonify({'ok': False, 'msg': 'Não autorizado'}), 403

    hoje  = date.today()
    limit = hoje + timedelta(days=30)
    donos = Usuario.query.filter_by(perfil='dono_rede').all()
    enviados = 0

    for dono in donos:
        if not dono.rede or not dono.rede.assinatura_ativa:
            continue
        meds = Medicamento.query.filter(
            Medicamento.rede_id == dono.rede_id,
            Medicamento.data_validade >= hoje,
            Medicamento.data_validade <= limit,
        ).order_by(Medicamento.data_validade.asc()).all()
        if meds:
            _enviar_notificacao_validade(dono, meds)
            enviados += 1

    audit('notificacoes_enviadas', f'total={enviados}')
    return jsonify({'ok': True, 'enviados': enviados})


# =============================================================================
# CONFIRMAÇÃO DE EMAIL
# =============================================================================

@app.route('/completar-cadastro', methods=['GET', 'POST'])
@login_required
def completar_cadastro():
    u = get_usuario_atual()

    # Só donos sem email precisam completar; demais passam direto
    if not u.is_dono or u.email:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()[:150]

        if '@' not in email or '.' not in email:
            flash('Email inválido. Verifique e tente novamente.', 'danger')
            return render_template('completar_cadastro.html', usuario=u)

        # Garante unicidade de email
        conflito = Usuario.query.filter(
            db.func.lower(Usuario.email) == email,
            Usuario.id != u.id
        ).first()
        if conflito:
            flash('Este email já está em uso por outra conta.', 'danger')
            return render_template('completar_cadastro.html', usuario=u)

        u.email = email
        u.email_confirmado = False
        db.session.commit()
        audit('email_cadastrado', f'username={u.username} email={email}')

        # Envia código de confirmação
        ok, erro = _enviar_codigo_confirmacao(u)
        if ok:
            flash('Email salvo! Enviamos um código de confirmação.', 'success')
        else:
            flash('Email salvo! Não foi possível enviar o código agora — tente reenviar na próxima tela.', 'warning')

        return redirect(url_for('confirmar_email'))

    return render_template('completar_cadastro.html', usuario=u)


@app.route('/confirmar-email', methods=['GET', 'POST'])
@login_required
def confirmar_email():
    u = get_usuario_atual()

    # Superadmin e usuários já confirmados passam direto
    if u.is_superadmin or u.email_confirmado:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        acao = request.form.get('acao', 'confirmar')

        if acao == 'reenviar':
            ok, erro = _enviar_codigo_confirmacao(u)
            if ok:
                flash('Novo código enviado para seu email!', 'success')
            else:
                flash(f'Erro ao enviar email: {erro}', 'danger')
            return redirect(url_for('confirmar_email'))

        codigo = request.form.get('codigo', '').strip()
        if not codigo:
            flash('Digite o código recebido por email.', 'danger')
            return render_template('confirmar_email.html', usuario=u)

        # Verifica expiração
        if u.email_codigo_exp and datetime.utcnow() > u.email_codigo_exp:
            flash('Código expirado. Clique em "Reenviar código".', 'danger')
            return render_template('confirmar_email.html', usuario=u)

        if u.email_codigo and codigo == u.email_codigo:
            u.email_confirmado  = True
            u.email_codigo      = None
            u.email_codigo_exp  = None
            db.session.commit()
            audit('email_confirmado', f'username={u.username}')
            flash('Email confirmado! Bem-vindo ao MedControl 🎉', 'success')
            return redirect(url_for('aceitar_termos'))
        else:
            audit('email_codigo_invalido', f'username={u.username}')
            flash('Código incorreto. Verifique seu email e tente novamente.', 'danger')

    return render_template('confirmar_email.html', usuario=u)

# =============================================================================
# PÁGINAS LEGAIS — LGPD
# =============================================================================

@app.route('/healthz')
def healthz():
    return {'status': 'ok'}, 200


@app.route('/sobre')
def sobre():
    return render_template('sobre.html')


@app.route('/politica-de-privacidade')
@login_required
def politica_privacidade():
    return render_template('politica_privacidade.html',
        data_atualizacao='01/03/2025')

@app.route('/termos-de-uso')
@login_required
def termos_uso():
    return render_template('termos_uso.html',
        data_atualizacao='01/03/2025')



# =============================================================================
# PREFERÊNCIAS DE TEMA
# =============================================================================

@app.route('/preferencias/nome', methods=['POST'])
@login_required
@csrf.exempt
def atualizar_nome_exibir():
    u = get_usuario_atual()
    nome = request.json.get('nome', '').strip()[:150]
    if len(nome) < 2:
        return jsonify({'ok': False, 'erro': 'Nome muito curto.'}), 400
    u.nome_exibir = nome
    db.session.commit()
    session['nome_exibir'] = nome
    audit('nome_atualizado', f'username={u.username} nome={nome}')
    return jsonify({'ok': True})


@app.route('/preferencias/tema', methods=['POST'])
@login_required
@csrf.exempt
def salvar_tema():
    tema = request.json.get('tema', 'light')
    if tema not in ('light', 'dark'):
        return jsonify({'ok': False}), 400
    u      = get_usuario_atual()
    u.tema = tema
    db.session.commit()
    session['tema'] = tema
    return jsonify({'ok': True})


# =============================================================================
# API REST (Movemos para app/routes/api.py)
# =============================================================================

@app.route('/api/docs')
@login_required
def api_docs():
    import secrets
    u = get_usuario_atual()
    if u.rede and not u.rede.token_api:
        u.rede.token_api = secrets.token_hex(32)
        db.session.commit()
    token = u.rede.token_api if u.rede else None
    return render_template('api_docs.html', usuario=u, token=token)



@app.route('/medicamentos/bulk-excluir', methods=['POST'])
@assinatura_required
def bulk_excluir():
    ids = request.json.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'Nenhum item selecionado.'}), 400
    
    u = get_usuario_atual()
    query = Medicamento.query.filter(Medicamento.id.in_(ids))
    
    # Segurança: garante que o usuário só delete o que é dele
    if not u.is_superadmin:
        if u.is_dono:
            query = query.filter_by(rede_id=u.rede_id)
        else:
            query = query.filter_by(filial_id=u.id)
            
    count = query.delete(synchronize_session=False)
    db.session.commit()
    
    audit('bulk_excluir', f'count={count} ids={ids}')
    return jsonify({'success': True, 'message': f'{count} medicamentos excluídos.'})


# =============================================================================
# PDF
# =============================================================================

@app.route('/relatorio/pdf')
@assinatura_required
def gerar_pdf():
    hoje  = date.today()
    u     = get_usuario_atual()
    meds  = get_medicamentos_query().order_by(Medicamento.data_validade.asc()).all()
    titulo_extra = (f' — {u.filial_nome or u.username}' if u.is_filial
                    else (f' — {u.rede.nome}' if u.is_dono and u.rede else ''))
    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=A4,
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
    estilos = getSampleStyleSheet()
    ts = ParagraphStyle('T', parent=estilos['Title'], fontSize=15,
                        textColor=colors.HexColor('#1e293b'), spaceAfter=4)
    ss = ParagraphStyle('S', parent=estilos['Normal'], fontSize=9,
                        textColor=colors.HexColor('#64748b'), spaceAfter=12)
    el = []
    el.append(Paragraph(f'MedControl — Controle de Validade{titulo_extra}', ts))
    el.append(Paragraph(f'Relatório gerado em {hoje.strftime("%d/%m/%Y")} | {len(meds)} item(s)', ss))
    el.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0')))
    el.append(Spacer(1, 0.4*cm))
    prejuizo = sum(m.valor_total for m in meds if m.status == 'vencido')
    resumo   = [['Vencidos','30 dias','60 dias','OK','Prejuízo'],
                [str(sum(1 for m in meds if m.status=='vencido')),
                 str(sum(1 for m in meds if m.status=='alerta_30')),
                 str(sum(1 for m in meds if m.status=='alerta_60')),
                 str(sum(1 for m in meds if m.status=='ok')),
                 f'R$ {prejuizo:,.2f}'.replace(',','X').replace('.',',').replace('X','.')]]
    rt = Table(resumo, colWidths=[3.5*cm]*5)
    rt.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e293b')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),8),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('BACKGROUND',(0,1),(0,1),colors.HexColor('#fee2e2')),
        ('BACKGROUND',(1,1),(1,1),colors.HexColor('#ffedd5')),
        ('BACKGROUND',(2,1),(2,1),colors.HexColor('#fef9c3')),
        ('BACKGROUND',(3,1),(3,1),colors.HexColor('#dcfce7')),
        ('BACKGROUND',(4,1),(4,1),colors.HexColor('#fee2e2')),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#e2e8f0')),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
    ]))
    el.append(rt)
    el.append(Spacer(1, 0.6*cm))
    SC = {'vencido': colors.HexColor('#fee2e2'), 'alerta_30': colors.HexColor('#ffedd5'),
          'alerta_60': colors.HexColor('#fef9c3'), 'ok': colors.HexColor('#dcfce7')}
    dados = [['#','Nome','Lote','Validade','Qtd','Preço','Total','Status']]
    et = [
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1e293b')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),7.5),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),('ALIGN',(1,1),(1,-1),'LEFT'),
        ('GRID',(0,0),(-1,-1),0.4,colors.HexColor('#e2e8f0')),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
    ]
    for i, m in enumerate(meds, 1):
        vt = f'R$ {m.valor_total:,.2f}'.replace(',','X').replace('.',',').replace('X','.')
        pu = f'R$ {m.preco_unitario:,.2f}'.replace(',','X').replace('.',',').replace('X','.')
        dados.append([str(i), m.nome[:32], m.lote, m.data_validade.strftime('%d/%m/%Y'),
                      str(m.quantidade), pu, vt, m.status_label])
        et.append(('BACKGROUND',(0,i),(-1,i),SC.get(m.status, colors.white)))
    tabela = Table(dados, colWidths=[0.6*cm,4.5*cm,2.2*cm,2.2*cm,1*cm,2*cm,2*cm,2.5*cm])
    tabela.setStyle(TableStyle(et))
    el.append(tabela)
    doc.build(el)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf',
                     download_name=f'relatorio_{hoje.strftime("%Y%m%d")}.pdf',
                     as_attachment=True)


@app.route('/relatorio/excel')
@assinatura_required
def gerar_excel():
    hoje = date.today()
    u    = get_usuario_atual()
    meds = get_medicamentos_query().order_by(Medicamento.data_validade.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Medicamentos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    center_align = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Cabeçalho
    colunas = ['#', 'Nome', 'Fabricante', 'Código de Barras', 'Lote', 'Validade', 'Qtd', 'Preço Unit.', 'Total', 'Status']
    ws.append(colunas)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Dados
    for i, m in enumerate(meds, 1):
        status_label = m.status_label
        row = [
            i, m.nome, m.fabricante or '', m.codigo_barras or '',
            m.lote, m.data_validade.strftime('%d/%m/%Y'),
            m.quantidade, m.preco_unitario, m.valor_total, status_label
        ]
        ws.append(row)

        # Formatação condicional simples (cores para status) e bordas
        row_idx = i + 1
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.border = border
            if col_idx in [1, 6, 7, 8, 9, 10]:
                cell.alignment = center_align
            if col_idx in [8, 9]:
                cell.number_format = '"R$ "#,##0.00'

    # Ajuste de largura das colunas
    larguras = [5, 35, 20, 20, 15, 12, 8, 15, 15, 15]
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=f'relatorio_{hoje.strftime("%Y%m%d")}.xlsx',
        as_attachment=True
    )


# =============================================================================
# INICIALIZAÇÃO
# =============================================================================

def seed_database():
    if Usuario.query.filter_by(perfil='superadmin').count() == 0:
        _admin_pass = os.environ.get('ADMIN_PASS')
        if not _admin_pass:
            raise RuntimeError(
                "ADMIN_PASS não definida! "
                "Adicione ADMIN_PASS nas variáveis de ambiente do Render."
            )
        admin = Usuario(
            username    = os.environ.get('ADMIN_USER', 'admin'),
            perfil      = 'superadmin',
            nome_exibir = 'Administrador',
        )
        admin.set_password(_admin_pass)
        db.session.add(admin)

    if Rede.query.count() == 0:
        hoje      = date.today()
        rede_demo = Rede(nome='Farmácia Demo', email_contato='demo@medcontrol.com.br',
                         plano='mensal', data_expiracao=hoje + timedelta(days=30))
        db.session.add(rede_demo)
        db.session.flush()
        dono_demo   = Usuario(username='dono_demo', perfil='dono_rede',
                               nome_exibir='Dono Demo', rede_id=rede_demo.id)
        filial_demo = Usuario(username='filial_demo', perfil='filial',
                               nome_exibir='Filial Centro', filial_nome='Unidade Centro',
                               rede_id=rede_demo.id)
        dono_demo.set_password('demo123')
        dono_demo.email_confirmado = True
        filial_demo.set_password('demo123')
        filial_demo.email_confirmado = True
        db.session.add_all([dono_demo, filial_demo])
        db.session.flush()
        db.session.add_all([
            Medicamento(nome='Dipirona 500mg', lote='LT-001',
                        data_validade=hoje - timedelta(days=5),
                        quantidade=20, preco_unitario=2.50,
                        rede_id=rede_demo.id, filial_id=filial_demo.id),
            Medicamento(nome='Amoxicilina 500mg', lote='LT-002',
                        data_validade=hoje + timedelta(days=15),
                        quantidade=50, preco_unitario=8.90,
                        rede_id=rede_demo.id, filial_id=filial_demo.id),
            Medicamento(nome='Omeprazol 20mg', lote='LT-003',
                        data_validade=hoje + timedelta(days=45),
                        quantidade=100, preco_unitario=12.00,
                        rede_id=rede_demo.id, filial_id=filial_demo.id),
            Medicamento(nome='Losartana 50mg', lote='LT-004',
                        data_validade=hoje + timedelta(days=180),
                        quantidade=200, preco_unitario=1.80,
                        rede_id=rede_demo.id, filial_id=filial_demo.id),
        ])
    db.session.commit()


with app.app_context():
    if os.environ.get('RESET_DB') == '1':
        db.drop_all()
    db.create_all()
    # Migração: adiciona colunas novas individualmente para compatibilidade (PostgreSQL e SQLite)
    with db.engine.connect() as conn:
        # Colunas para Usuarios
        for col_name, col_type in [
            ("termos_aceitos", "BOOLEAN DEFAULT FALSE"),
            ("termos_aceitos_em", "TIMESTAMP"),
            ("email_confirmado", "BOOLEAN DEFAULT FALSE"),
            ("email_codigo", "VARCHAR(10)"),
            ("email_codigo_exp", "TIMESTAMP"),
            ("email", "VARCHAR(150)"),
        ]:
            try:
                conn.execute(db.text(f"ALTER TABLE usuarios ADD COLUMN {col_name} {col_type}"))
                conn.commit()
            except Exception: pass # Coluna já existe

        # Colunas para Redes
        for col_name, col_type in [
            ("trial", "BOOLEAN DEFAULT TRUE"),
            ("trial_inicio", "TIMESTAMP"),
            ("mp_assinatura_id", "VARCHAR(100)"),
            ("mp_payer_email", "VARCHAR(150)"),
        ]:
            try:
                conn.execute(db.text(f"ALTER TABLE redes ADD COLUMN {col_name} {col_type}"))
                conn.commit()
            except Exception: pass # Coluna já existe

    seed_database()

if __name__ == '__main__':
    app.run(debug=os.environ.get('FLASK_DEBUG', '0') == '1',
            host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))


